import pywhatkit
from openpyxl.workbook import Workbook
from openpyxl import load_workbook
from pynput.keyboard import Key,Controller

wb = load_workbook('Book.xlsx')
ws = wb.active
t = 2
start = 0
end = 3
keyb = Controller()
    
for i in range(start,end):
    ph = f'{ws["A"+str(t)].value}'
    ms = f'{ws["B"+str(t)].value}'
    phone = "91+" + ph
    msg = "Hello " + ms + " Namaste"
    print(phone)
    print(msg) 
    pywhatkit.sendwhatmsg_instantly(phone, msg)
    t = t+1