from distutils import core
from tkinter import *
from tkinter import messagebox
import openpyxl as xl;
from tkinter import filedialog
import pywhatkit
from openpyxl.workbook import Workbook
from openpyxl import load_workbook
import time
import pyautogui
from pynput.keyboard import Key,Controller
import webbrowser as web
from datetime import datetime
from typing import Optional
from urllib.parse import quote
import pyautogui as pg
from pywhatkit.core import core, exceptions, log

master=Tk()

header_l = Label(master,text="WHATSAPP MSG SENDER",font= ('Helvetica 24 underline'))
header_l.grid(row=1)
header_l.place(relx=0.5,y=35,anchor=CENTER)

#File
def opennew():
    #copy
    master.filename = filedialog.askopenfilename()
    wb1 = xl.load_workbook(master.filename) 
    ws1 = wb1.worksheets[0]

    #paste
    file2name = "finalsheet.xlsx"
    wb2 = xl.load_workbook(file2name) 
    ws2 = wb2.active
    mr=ws1.max_row
    mc=ws1.max_column

    for i in range (1,mr+1):
        for j in range (1,mc+1):
            c=ws1.cell(row=i,column=j)

            ws2.cell(row=i,column=j).value=c.value

    #Save
    wb2.save(str(file2name))

file_l = Label(master,text="Select File")
file_l.grid(row=2)
file_l.place(relx=0.5,y=190,anchor=CENTER)

file_btn = Button(master,text="Open",command=opennew,width=30)
file_btn.grid(row=3)
file_btn.place(relx=0.5,y=220,anchor=CENTER)

#Text (Start & End)
def print_input():
    print(start_t.get() + " " + end_t.get() + " " + date_t.get())

lb_st = Label(master,text="From Where To Start")
lb_st.grid(row=4)
lb_st.place(x=110,y=100,anchor=CENTER)

start_t = Entry(master,width=20)
start_t.grid(row=5)
start_t.place(x=110,y=130,anchor=CENTER)

lb_en = Label(master,text="End")
lb_en.grid(row=6)
lb_en.place(x=570,y=100,anchor=CENTER)

end_t = Entry(master,width=20)
end_t.grid(row=7)
end_t.place(x=570,y=130,anchor=CENTER)

#Text (Date)
lb_dt = Label(master,text="Enter The Date")
lb_dt.grid(row=8)
lb_dt.place(x=120,y=290,anchor=CENTER)

date_t = Entry(master,width=30)
date_t.grid(row=9)
date_t.place(x=130,y=320,anchor=CENTER)

#text_btn = Button(master,text="Print",command=print_input).grid(row=10)

#Customize Msg (Optional)
lb_cm = Label(master,text="Enter The Msg(Optional)")
lb_cm.grid(row=10)
lb_cm.place(x=570,y=290,anchor=CENTER)

cm_t = Entry(master,width=30)
cm_t.grid(row=12,sticky=W)
cm_t.place(x=560,y=320,anchor=CENTER)

#Sending Msg
def main_f():
    wb = load_workbook("finalsheet.xlsx")
    ws = wb.active
    start = int(start_t.get())
    end = int(end_t.get())+1
    keyb = Controller()
        
    for i in range(start,end):
        ph = f'{ws["A"+str(start)].value}'
        ms = f'{ws["B"+str(start)].value}'
        phone = "91+" + ph
        msg = "Dear Customer, your EMI to Bazaari Global Finance of Rs " + ms + " is due on " + date_t.get() + " and you are requested to maintain adequate balance on " + date_t.get() + " onwards. For any clarification you may reach us on 91-8452816111"
        print(phone , " " , end)
        print(msg) 
        #pywhatkit.sendwhatmsg_instantly(phone, msg)
        sendwhatmsg_instantly(phone, msg)
        time.sleep(2)
        close_tab()
        start = start+1
    last_msg()

final_btn = Button(master,text="BEGIN",command=main_f,width=30)
final_btn.grid(row=13)
final_btn.place(relx=0.5,y=420,anchor=CENTER)

#Alert Box
def last_msg():
   messagebox.showinfo("Done", "All Sent")

#Close Tab
def close_tab():
    x=198
    y=31
    a=288
    b=18

    pyautogui.moveTo(x,y)
    pyautogui.click()

    pyautogui.moveTo(a,b)
    pyautogui.click()

#send msg
def sendwhatmsg_instantly(
    phone_no: str,
    message: str,
    wait_time: int = 15,
    tab_close: bool = False,
    close_time: int = 3,
) -> None:
    """Send WhatsApp Message Instantly"""

    if not core.check_number(number=phone_no):
        raise exceptions.CountryCodeException("Country Code Missing in Phone Number!")

    web.open(f"https://web.whatsapp.com/send?phone={phone_no}&text={quote(message)}")
    time.sleep(4)
    pg.click(core.WIDTH / 2, core.HEIGHT / 2)
    time.sleep(10)
    pg.press("enter")
    if tab_close:
        core.close_tab(wait_time=close_time)

master.title("Whatsapp Msg Sender")
master.geometry("700x500")
master.mainloop()