from contextlib import nullcontext
from distutils import core
from tkinter import *
from tkinter import messagebox
import openpyxl as xl;
from tkinter import filedialog
from openpyxl.workbook import Workbook
from openpyxl import load_workbook
import time
import pyautogui
from pynput.keyboard import Controller
import webbrowser as web
from urllib.parse import quote
import pyautogui as pg
from pywhatkit.core import core, exceptions

master=Tk()

header_l = Label(master,text="WHATSAPP MSG SENDER",font= ('Helvetica 24 underline'),bg='green')
header_l.grid(row=1)
header_l.place(relx=0.5,y=35,anchor=CENTER)

#File
def opennew():
    #copy
    master.filename = filedialog.askopenfilename()
    wb1 = xl.load_workbook(master.filename) 
    ws1 = wb1.worksheets[0]

    #paste
    file2name = "finalsheet.xlsx"
    wb2 = xl.load_workbook(file2name) 
    ws2 = wb2.active
    mr=ws1.max_row
    mc=ws1.max_column

    for i in range (1,mr+1):
        for j in range (1,mc+1):
            c=ws1.cell(row=i,column=j)

            ws2.cell(row=i,column=j).value=c.value

    #Save
    wb2.save(str(file2name))

file_l = Label(master,text="Select A File",bg='green')
file_l.grid(row=2)
file_l.place(relx=0.5,y=190,anchor=CENTER)

file_btn = Button(master,text="Open",command=opennew,width=20,bg='light blue')
file_btn.grid(row=3)
file_btn.place(relx=0.5,y=230,anchor=CENTER)

#Text (Start & End)
def print_input():
    print(start_t.get() + " " + end_t.get() + " " + dt_fi.get())

lb_st = Label(master,text="From Where To Start",bg='green')
lb_st.grid(row=4)
lb_st.place(x=110,y=100,anchor=CENTER)

start_t = Entry(master,width=20,bg='light blue')
start_t.grid(row=5)
start_t.place(x=110,y=130,anchor=CENTER)

lb_en = Label(master,text="End",bg='green')
lb_en.grid(row=6)
lb_en.place(x=570,y=100,anchor=CENTER)

end_t = Entry(master,width=20,bg='light blue')
end_t.grid(row=7)
end_t.place(x=570,y=130,anchor=CENTER)

#Column A
columA = Label(master,text="Enter 1st Field",bg='green')
columA.grid(row=6)
columA.place(x=110,y=190,anchor=CENTER)

columA_t = Entry(master,width=15,bg='light blue')
columA_t.grid(row=6)
columA_t.place(x=110,y=220,anchor=CENTER)

columB = Label(master,text="Enter 2nd Field",bg='green')
columB.grid(row=6)
columB.place(x=565,y=190,anchor=CENTER)

columB_t = Entry(master,width=15,bg='light blue')
columB_t.grid(row=6)
columB_t.place(x=565,y=220,anchor=CENTER)

#Text (Date)
lb_dt = Label(master,text="Enter The Date",bg='green')
lb_dt.grid(row=8)
lb_dt.place(x=120,y=290,anchor=CENTER)

dt_fi = Entry(master,width=30,bg='light blue')
dt_fi.grid(row=9)
dt_fi.place(x=130,y=320,anchor=CENTER)

#text_btn = Button(master,text="Print",command=print_input).grid(row=10)

#Customize Msg (Optional)
lb_cm = Label(master,text="Enter The Msg(Optional)",bg='green')
lb_cm.grid(row=10)
lb_cm.place(x=570,y=290,anchor=CENTER)

cm_t = Entry(master,width=30,bg='light blue')
cm_t.grid(row=12,sticky=W)
cm_t.place(x=560,y=320,anchor=CENTER)

#Sending Msg
def main_f():
    wb = load_workbook("finalsheet.xlsx")
    ws = wb.active
    start = int(start_t.get())
    end = int(end_t.get())+1
    keyb = Controller()
    A_col = columA_t.get()
    B_col = columB_t.get()

    for i in range(start,end):
        ph = f'{ws[A_col+str(start)].value}'
        ms = f'{ws[B_col+str(start)].value}'
        phone = "91+" + ph
        msg = se_msg(phone,ms)
    
        if(dt_fi==nullcontext):
            date_msg = ""
        else:
            date_msg = dt_fi.get()

        if(len(msg)==0):
            msg = "Dear Customer, your EMI to Bazaari Global Finance of " + ms + " is due on " + date_msg + " and you are requested to maintain adequate balance on " + dt_fi.get() + " onwards. For any clarification you may reach us on 91-8452816111"
        
        print(phone)
        print(msg) 
        #sendwhatmsg_instantly(phone, msg)
        time.sleep(2)
        close_tab()
        start = start+1
    close_tab()    
    last_msg()

final_btn = Button(master,text="BEGIN",command=main_f,width=30,bg='light blue')
final_btn.grid(row=13)
final_btn.place(relx=0.5,y=450,anchor=CENTER)

#Alert Box
def last_msg():
   messagebox.showinfo("Done", "All Sent")

#Close Tab
def close_tab():
    x=198
    y=31
    a=288
    b=18

    pyautogui.moveTo(x,y)
    pyautogui.click()

    pyautogui.moveTo(a,b)
    pyautogui.click()

#send msg
def sendwhatmsg_instantly(
    phone_no: str,
    message: str,
    wait_time: int = 15,
    tab_close: bool = False,
    close_time: int = 3,) ->None:
    """Send WhatsApp Message Instantly"""

    if not core.check_number(number=phone_no):
        raise exceptions.CountryCodeException("Country Code Missing in Phone Number!")

    web.open(f"https://web.whatsapp.com/send?phone={phone_no}&text={quote(message)}")
    time.sleep(4)
    pg.click(core.WIDTH / 2, core.HEIGHT / 2)
    time.sleep(10)
    pg.press("enter")
    if tab_close:
        core.close_tab(wait_time=close_time)

def se_msg(ms,ms1):
    a = cm_t.get()
    flag = 0
    flag1 = 0
    A_posi = 0
    final_msg = ""
    Y_posi = 0
    for i in range(1,len(a)):
        if(a[i]=='A'):
            if(a[i+1]=='A'):
                if(a[i+2]=='A'):
                    if(a[i+3]=='A'):
                        A_posi = i
                        flag = 1

        elif(a[i]=='Y'): 
            if(a[i+1]=='Y'):
                if(a[i+2]=='Y'):
                    if(a[i+3]=='Y'):
                        Y_posi = i
                        flag = 1
                        flag1 = 1
                        
    if(flag == 1):
        A_posiend = A_posi + 3
        Y_posiend = Y_posi + 3
        msg = ""
        msg1 = ""

        for i in range(0,A_posi):
            msg += a[i] 

        final_msg = msg + ms

        if(flag1==1):
            for i in range(A_posiend,Y_posi):
                final_msg += a[i]   

            for i in range(Y_posiend+1,len(a)):
                msg1 += a[i]

            final_msg = final_msg + ms1 + msg1   

        else:
            for i in range(A_posiend+1,len(a)):
                msg1 += a[i]
            final_msg = final_msg + msg1

        return final_msg
    else:
        return a

master.title("Whatapp Bulk Msg Sender")
master.iconbitmap("whatsapp.ico")
master.geometry("700x500")
master.configure(bg='green')
master.mainloop()